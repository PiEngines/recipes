"""Import script for the Bundeslebensmittelschlüssel (BLS) 4.0 ingredient database.

Downloads the BLS 4.0 dataset, filters it down to base ingredients (excluding
prepared dishes/recipes and cooked/processed forms), and writes the result to
backend/app/data/bls_ingredients.json.

# Datenquelle: Bundeslebensmittelschlüssel (BLS) 4.0, Max Rubner-Institut, CC BY 4.0

Usage:
    python scripts/import_bls.py
"""
import html
import io
import json
import re
import zipfile
from pathlib import Path

import httpx
from openpyxl import load_workbook

DOWNLOAD_PAGE = "https://www.blsdb.de/download"
DATA_DIR = Path(__file__).resolve().parent.parent / "app" / "data"
OUTPUT_PATH = DATA_DIR / "bls_ingredients.json"

# BLS codes starting with these prefixes are prepared dishes/recipes, not base ingredients
EXCLUDED_CODE_PREFIXES = ("X", "Y")

# Name keywords indicating a cooked/processed form rather than a raw base ingredient
PREPARATION_KEYWORDS = (
    "gegart", "gekocht", "gebraten", "gedünstet", "tiefgefroren", "getrocknet",
)

# Column indices (0-based) in the "BLS_4_0_Daten_*_DE.xlsx" data sheet
COL_BLS_CODE = 0
COL_NAME_DE = 1
COL_ENERGY_KCAL = 6
COL_PROTEIN_G = 12
COL_FAT_G = 15
COL_CARBS_G = 18
COL_FIBER_G = 21


def find_xlsx_download_url() -> str:
    """The /download page is a landing page with a tokenized link to the data zip."""
    resp = httpx.get(DOWNLOAD_PAGE, follow_redirects=True, timeout=30)
    resp.raise_for_status()
    page = html.unescape(resp.text)
    match = re.search(r'href="(/assets/uploads/BLS_[^"]+\.zip\?token=[^"]+)"', page)
    if not match:
        raise RuntimeError("Could not find BLS download link on landing page")
    return f"https://www.blsdb.de{match.group(1)}"


def download_data_workbook(zip_url: str):
    resp = httpx.get(zip_url, follow_redirects=True, timeout=120)
    resp.raise_for_status()
    with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
        data_file = next(n for n in zf.namelist() if re.search(r"Daten.*\.xlsx$", n))
        with zf.open(data_file) as f:
            return load_workbook(io.BytesIO(f.read()), read_only=True)


def to_float_or_none(value):
    if isinstance(value, (int, float)):
        return float(value)
    return None  # covers missing values and markers like "TR", "<LOQ", "<LOD", "-"


def is_base_ingredient(bls_code: str, name: str) -> bool:
    if bls_code[:1] in EXCLUDED_CODE_PREFIXES:
        return False
    name_lower = name.lower()
    return not any(keyword in name_lower for keyword in PREPARATION_KEYWORDS)


def main():
    print(f"Fetching download link from {DOWNLOAD_PAGE} ...")
    zip_url = find_xlsx_download_url()

    print("Downloading and extracting BLS data workbook ...")
    wb = download_data_workbook(zip_url)
    ws = wb[wb.sheetnames[0]]

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        bls_code = row[COL_BLS_CODE]
        name = row[COL_NAME_DE]
        if not bls_code or not name:
            continue
        if not is_base_ingredient(bls_code, name):
            continue
        items.append({
            "bls_id": bls_code,
            "name": name,
            "energie_kcal": to_float_or_none(row[COL_ENERGY_KCAL]),
            "protein_g": to_float_or_none(row[COL_PROTEIN_G]),
            "fett_g": to_float_or_none(row[COL_FAT_G]),
            "kohlenhydrate_g": to_float_or_none(row[COL_CARBS_G]),
            "ballaststoffe_g": to_float_or_none(row[COL_FIBER_G]),
        })

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=2)

    print(f"Wrote {len(items)} ingredients to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
